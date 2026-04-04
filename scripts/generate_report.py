#!/usr/bin/env python3
"""
generate_report.py — Génère un rapport Excel (.xlsx) depuis les résultats de pré-évaluation Epitech.

Usage:
    python3 generate_report.py --results /tmp/eval-results.json --output /tmp/pre-eval-projet-date.xlsx

Structure JSON attendue (--results):
{
  "project": "Nom du projet",
  "student": "Prénom Nom (optionnel)",
  "date": "2024-01-15",
  "criteria_results": [
    {
      "id": "C1",
      "label": "Description du critère",
      "category": "fonctionnel",
      "points_max": 10,
      "status": "validated",   // validated | partial | failed | blocking
      "points_obtained": 10,
      "remarks": "Implémentation trouvée dans src/auth/login.ts"
    }
  ],
  "bad_practices": [
    {
      "file": "src/config.ts",
      "line": 12,
      "problem": "API key en dur",
      "severity": "critical"   // critical | medium | minor
    }
  ]
}
"""

import json
import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl non installé. Lancer : pip install openpyxl")
    sys.exit(1)


# ─── Couleurs ────────────────────────────────────────────────────────────────

COLORS = {
    "validated":  "C8E6C9",  # vert clair
    "partial":    "FFF9C4",  # jaune clair
    "failed":     "FFCDD2",  # rouge clair
    "blocking":   "E1BEE7",  # violet clair
    "critical":   "FFCDD2",
    "medium":     "FFF9C4",
    "minor":      "E3F2FD",  # bleu très clair
    "header":     "1565C0",  # bleu foncé
    "summary_bg": "E8EAF6",  # violet très clair
}

STATUS_LABELS = {
    "validated": "✅ VALIDÉ",
    "partial":   "⚠️ PARTIEL",
    "failed":    "❌ NON VALIDÉ",
    "blocking":  "🚫 BLOQUANT",
}

SEVERITY_LABELS = {
    "critical": "🔴 Critique",
    "medium":   "🟡 Moyen",
    "minor":    "🟢 Mineur",
}

POINTS_MAP = {
    "validated": lambda r: r.get("points_max", 0),
    "partial":   lambda r: round(r.get("points_max", 0) * 0.5, 1),
    "failed":    lambda r: 0,
    "blocking":  lambda r: 0,
}

ESTIMATE_MARGIN = 15


def compute_estimate_range(points_obtained: float, points_max: float) -> Tuple[float, float]:
    """Calcule la fourchette estimée ±ESTIMATE_MARGIN, avec bornes [0, points_max].

    points_obtained est d'abord borné à [0, points_max], ce qui garantit
    pts_min <= points_obtained <= pts_max_est sans swap supplémentaire.
    """
    points_max = max(0, points_max)
    points_obtained = min(max(0, points_obtained), points_max)
    pts_min = max(0, points_obtained - ESTIMATE_MARGIN)
    pts_max_est = min(points_max, points_obtained + ESTIMATE_MARGIN)
    return pts_min, pts_max_est


def format_points(value: float) -> str:
    """Formate un nombre de points : entier si .0, sinon 1 décimale."""
    formatted = f"{value:.1f}"
    return formatted[:-2] if formatted.endswith(".0") else formatted


def compute_status_counts(criteria_results: List[dict]) -> Dict[str, int]:
    """Retourne le comptage des statuts pour une liste de critères."""
    counts: Dict[str, int] = {"validated": 0, "partial": 0, "failed": 0, "blocking": 0}
    for c in criteria_results:
        status = c.get("status", "failed")
        counts[status] = counts.get(status, 0) + 1
    return counts


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def make_header_font(color: str = "FFFFFF") -> Font:
    return Font(bold=True, color=color)


def thin_border() -> Border:
    side = Side(style="thin", color="BDBDBD")
    return Border(left=side, right=side, top=side, bottom=side)


def set_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws, headers: list, row: int = 1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = make_fill(COLORS["header"])
        cell.font = make_header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


# ─── Onglet Résultats ─────────────────────────────────────────────────────────

def build_results_sheet(ws, criteria_results: list):
    ws.title = "Résultats"
    ws.row_dimensions[1].height = 30

    headers = ["ID", "Critère", "Catégorie", "Points max", "Statut", "Points obtenus", "Remarques"]
    write_header_row(ws, headers)

    totals = {"points_max": 0, "points_obtained": 0}

    for row_idx, criterion in enumerate(criteria_results, start=2):
        status = criterion.get("status", "failed")
        points_max = criterion.get("points_max", 0)
        points_obtained = criterion.get("points_obtained") or POINTS_MAP.get(status, lambda r: 0)(criterion)

        row_fill = make_fill(COLORS.get(status, "FFFFFF"))
        values = [
            criterion.get("id", ""),
            criterion.get("label", ""),
            criterion.get("category", ""),
            points_max,
            STATUS_LABELS.get(status, status),
            points_obtained,
            criterion.get("remarks", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

        totals["points_max"] += points_max
        totals["points_obtained"] += points_obtained

    # Ligne totaux
    total_row = len(criteria_results) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=totals["points_max"]).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=totals["points_obtained"]).font = Font(bold=True)
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = make_fill(COLORS["summary_bg"])
        ws.cell(row=total_row, column=col).border = thin_border()

    set_column_widths(ws, {"A": 8, "B": 45, "C": 15, "D": 12, "E": 18, "F": 16, "G": 55})
    ws.freeze_panes = "A2"

    return totals


# ─── Onglet Mauvaises pratiques ───────────────────────────────────────────────

def build_bad_practices_sheet(ws, bad_practices: list):
    ws.title = "Mauvaises pratiques"
    ws.row_dimensions[1].height = 30

    if not bad_practices:
        ws.cell(row=2, column=1, value="✅ Aucune mauvaise pratique détectée.")
        return

    headers = ["Fichier", "Ligne", "Problème", "Sévérité"]
    write_header_row(ws, headers)

    for row_idx, practice in enumerate(bad_practices, start=2):
        severity = practice.get("severity", "minor")
        row_fill = make_fill(COLORS.get(severity, "FFFFFF"))
        values = [
            practice.get("file", ""),
            practice.get("line", ""),
            practice.get("problem", ""),
            SEVERITY_LABELS.get(severity, severity),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

    set_column_widths(ws, {"A": 40, "B": 8, "C": 55, "D": 15})
    ws.freeze_panes = "A2"


# ─── Onglet Récapitulatif ─────────────────────────────────────────────────────

def build_summary_sheet(ws, data: dict, totals: dict):
    ws.title = "Récapitulatif"

    criteria_results = data.get("criteria_results", [])

    counts = compute_status_counts(criteria_results)
    blocking_list = []

    for c in criteria_results:
        if c.get("status") == "blocking":
            blocking_list.append(f"[{c.get('id','')}] {c.get('label','')}")

    bg = make_fill(COLORS["summary_bg"])
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    def write_kv(row, key, value, value_color=None):
        k_cell = ws.cell(row=row, column=1, value=key)
        k_cell.font = bold
        k_cell.fill = bg
        k_cell.border = thin_border()
        v_cell = ws.cell(row=row, column=2, value=value)
        v_cell.fill = make_fill(value_color) if value_color else bg
        v_cell.border = thin_border()
        v_cell.alignment = center

    row = 1
    ws.cell(row=row, column=1, value="📋 Récapitulatif de pré-évaluation").font = Font(bold=True, size=14, color="1565C0")
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    write_kv(row, "Projet", data.get("project", "")); row += 1
    write_kv(row, "Étudiant", data.get("student", "Non renseigné")); row += 1
    write_kv(row, "Date", data.get("date", datetime.now().strftime("%Y-%m-%d"))); row += 1
    row += 1

    write_kv(row, "⚠️ Avertissement", "Ce rapport détecte des patterns, pas l'exécution."); row += 1
    row += 1

    pts_min, pts_max_est = compute_estimate_range(totals["points_obtained"], totals["points_max"])
    write_kv(row, "Fourchette estimée", f"{format_points(pts_min)}–{format_points(pts_max_est)} pts (±{ESTIMATE_MARGIN} selon vérification manuelle)"); row += 1
    row += 1

    ws.cell(row=row, column=1, value="Statuts").font = bold
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    write_kv(row, "✅ Validés", counts["validated"], COLORS["validated"]); row += 1
    write_kv(row, "⚠️ Partiels", counts["partial"], COLORS["partial"]); row += 1
    write_kv(row, "❌ Non validés", counts["failed"], COLORS["failed"]); row += 1
    write_kv(row, "🚫 Bloquants", counts["blocking"], COLORS["blocking"]); row += 1
    row += 1

    if blocking_list:
        ws.cell(row=row, column=1, value="🔴 Critères bloquants").font = Font(bold=True, color="C62828")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1
        for item in blocking_list:
            cell = ws.cell(row=row, column=1, value=item)
            cell.fill = make_fill(COLORS["failed"])
            cell.border = thin_border()
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

    set_column_widths(ws, {"A": 35, "B": 25})


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Génère un rapport Excel de pré-évaluation Epitech")
    parser.add_argument("--results", required=True, help="Chemin vers le fichier JSON des résultats")
    parser.add_argument("--output", required=True, help="Chemin du fichier Excel de sortie (.xlsx)")
    args = parser.parse_args()

    results_path = Path(args.results)
    if not results_path.exists():
        print(f"❌ Fichier introuvable : {results_path}")
        sys.exit(1)

    with open(results_path, encoding="utf-8") as f:
        data = json.load(f)

    criteria_results = data.get("criteria_results", [])
    bad_practices = data.get("bad_practices", [])

    if not criteria_results:
        print("⚠️  Aucun critère dans le JSON. Vérifier la structure du fichier.")

    wb = Workbook()
    ws_results = wb.active
    ws_bad = wb.create_sheet()
    ws_summary = wb.create_sheet()

    totals = build_results_sheet(ws_results, criteria_results)
    build_bad_practices_sheet(ws_bad, bad_practices)
    build_summary_sheet(ws_summary, data, totals)

    # Réordonner : Récapitulatif en premier
    wb.move_sheet(ws_summary, offset=-(len(wb.worksheets) - 1))

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Rapport Excel généré : {output_path}")
    counts = compute_status_counts(criteria_results)
    print(f"   Statuts : {counts['validated']} ✅ — {counts['partial']} ⚠️ — {counts['failed']} ❌ — {counts['blocking']} 🚫")
    pts_min, pts_max_est = compute_estimate_range(totals["points_obtained"], totals["points_max"])
    print(f"   Fourchette estimée : {format_points(pts_min)}–{format_points(pts_max_est)} pts (±{ESTIMATE_MARGIN} selon vérification manuelle)")
    print("   ⚠️ Ce rapport détecte des patterns, pas l'exécution.")


if __name__ == "__main__":
    main()
